from django.core.management.base import BaseCommand
from API.models import Hospitals
import openpyxl 
from django.db import connection, transaction



    
class Command(BaseCommand):
    help = 'Load hospitals from a text file'
    def reset_sequence(self):
        """Reset the auto-increment counter for the Hospitals table."""
        with connection.cursor() as cursor:
            table_name = Hospitals._meta.db_table
            if connection.vendor == 'postgresql':
                cursor.execute(
                    f"ALTER SEQUENCE {table_name}_id_seq RESTART WITH 1;"
                )
            elif connection.vendor == 'sqlite':
                cursor.execute(
                    f"UPDATE sqlite_sequence SET seq = 0 WHERE name = '{table_name}';"
                )
            elif connection.vendor == 'mysql':
                cursor.execute(
                    f"ALTER TABLE {table_name} AUTO_INCREMENT = 1;"
                )
        self.stdout.write("Reset database ID sequence to 1.")
        
    def load_data(self):
        file_path ='D:/projects/insurance/backend/djangoProject/API/management/commands/insurances.xlsx'
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            self.stdout.write(f'Total rows: {ws.max_row}, Columns: {ws.max_column}')
            for i in range(2,ws.max_row+1):
                values = [ws.cell(row=i, column=j).value for j in range(1,ws.max_column+1)]
                Hospitals.objects.create(insurance = values[0],name = values[1], city = values[2] , address = values[3] , phone = values[4] , services = values[5])
            self.stdout.write(self.style.SUCCESS('Successfully loaded hospitals!'))
            
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error: {str(e)}"))
    
    def handle(self, *args, **options):
        Hospitals.objects.all().delete()
        self.stdout.write("Deleted all existing hospital entries.")

        # Reset the database sequence for the ID field
        self.reset_sequence()

        # Load new data from Excel
        self.load_data()
            
        # Hospitals.objects.all().delete()
        # with open('../hospitals.txt', 'r' , encoding="utf_8") as f:
        #     for line in f:
        #         name = line.strip()
        #         Hospitals.objects.create(name=name)
        
    